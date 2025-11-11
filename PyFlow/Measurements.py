import csv
import matplotlib.pyplot as plt
import os
import numpy as np
import pickle

from openpyxl import load_workbook
from datetime import datetime
from scipy.interpolate import interp1d

def loadFromFile(filePath, delimiter=',', headerRow=1, startDataRow=2, headerRenameToMapping=None, headerParseFunctionMapping=None, sheetName=None):
    data = {}
    ext = os.path.splitext(filePath)[1].lower()

    # --- CSV handling ---
    if ext == ".csv":
        with open(filePath, newline="") as f:
            lines = f.readlines()

        header = lines[headerRow - 1].strip().split(delimiter)
        data_lines = lines[startDataRow - 1:]

        if headerRenameToMapping:
            header = [headerRenameToMapping.get(col, col) for col in header]

        reader = csv.DictReader(data_lines, fieldnames=header, delimiter=delimiter)
        rows = list(reader)

    # --- XLSX handling ---
    elif ext in [".xlsx", ".xlsm"]:
        wb = load_workbook(filePath, data_only=True)
        ws = wb[sheetName] if sheetName else wb.active

        rows = list(ws.iter_rows(values_only=True))
        header = list(rows[headerRow - 1])
        data_rows = rows[startDataRow - 1:]

        if headerRenameToMapping:
            header = [headerRenameToMapping.get(col, col) for col in header]

        rows = [dict(zip(header, row)) for row in data_rows]

    else:
        raise ValueError(f"Unsupported file type: {ext}")

    # --- Parse and build dictionary ---
    for row in rows:
        for key, value in row.items():
            parsed_value = value
            if headerParseFunctionMapping and key in headerParseFunctionMapping:
                try:
                    if isinstance(value, str):
                        parsed_value = headerParseFunctionMapping[key](value)
                    else:
                        # Parsed value seems to be valid
                        pass

                except Exception as e:
                    #print(f"Error parsing value '{value}' for column '{key}': {e}")
                    parsed_value = float('nan')

            data.setdefault(str(key).strip(), []).append(parsed_value)

    return data

def filterDataByKey(data, key, filterFunction):
    if key not in data: raise KeyError(f"Key '{key}' not found in data.")

    filteredData = {k: [] for k in data}
    for i, v in enumerate(data[key]):
        if filterFunction(v):
            for k in data:
                filteredData[k].append(data[k][i])

    return filteredData

def saveToPickle(data, filePath):
    with open(filePath, 'wb') as f:
        pickle.dump(data, f)

def loadFromPickle(filePath):
    with open(filePath, 'rb') as f:
        data = pickle.load(f)
    return data

def removeNaNEntries(data, keys=None):
    if keys is None: keys = data.keys()

    validIndices = [i for i in range(len(next(iter(data.values())))) if all(data[key][i] is not None and not (isinstance(data[key][i], float) and np.isnan(data[key][i])) for key in keys)]

    cleanedData = {key: [data[key][i] for i in validIndices] for key in data}
    return cleanedData

def reduceDataToTimeRange(data, timeKey, startTime, endTime):
    if timeKey not in data: raise KeyError(f"Time key '{timeKey}' not found in data.")

    reducedData = {key: [] for key in data}
    for i, t in enumerate(data[timeKey]):
        if startTime <= t <= endTime:
            for key in data:
                reducedData[key].append(data[key][i])

    return reducedData

def addRelativeTime(data, timeKey, relTimeKey="RelTime", timeOffsetInSeconds=0):
    if timeKey not in data: raise KeyError(f"Time key '{timeKey}' not found in data.")

    startTime = data[timeKey][0]
    data[relTimeKey] = [(t - startTime).total_seconds() + timeOffsetInSeconds if t is not None else None for t in data[timeKey]]
    return data

def addTimeOffset(data, timeKey_seconds, timeOffsetInSeconds):
    if timeKey_seconds not in data: raise KeyError(f"Time key '{timeKey_seconds}' not found in data.")

    data[timeKey_seconds] = [(t + timeOffsetInSeconds) if t is not None else None for t in data[timeKey_seconds]]
    return data

def mergeDataSetsOnTimeKey(dataSet1, dataSet2, timeKey, newCommonTimeVector=None):
    mergedData = {}

    if newCommonTimeVector is None:
        time1 = dataSet1[timeKey]
        time2 = dataSet2[timeKey]

        newCommonTimeVector = time1 
        if len(time1) > len(time2): newCommonTimeVector = time2

    mergedData[timeKey] = newCommonTimeVector
    interpolate = lambda data, time: interp1d(time, data, fill_value="extrapolate")(newCommonTimeVector).tolist()

    for key in set(dataSet1.keys()).union(set(dataSet2.keys())):
        if key == timeKey: continue

        if key in dataSet1 and key not in dataSet2: 
            try:
                mergedData[key] = interpolate(dataSet1[key], dataSet1[timeKey])
            except Exception as e:
                print(f"Error interpolating key '{key}' from dataSet1: {e}")
                mergedData[key] = [None] * len(newCommonTimeVector)

        elif key in dataSet2 and key not in dataSet1:
            try:
                mergedData[key] = interpolate(dataSet2[key], dataSet2[timeKey])
            except Exception as e:
                print(f"Error interpolating key '{key}' from dataSet2: {e}")
                mergedData[key] = [None] * len(newCommonTimeVector)

        else:
            try:
                data1Interp = interpolate(dataSet1[key], dataSet1[timeKey])
                data2Interp = interpolate(dataSet2[key], dataSet2[timeKey])
                mergedData[key] = [data2Interp[idx] if time >= dataSet2[timeKey][0] else data1Interp[idx] for idx, time in enumerate(newCommonTimeVector)]
            except Exception as e:
                print(f"Error interpolating key '{key}' from both data sets: {e}")
                mergedData[key] = [None] * len(newCommonTimeVector)

    return mergedData

def deriveNewKeyFromData(data, newKey, function):
    data[newKey] = [function({k: data[k][i] for k in data}) for i in range(len(data[next(iter(data))]))]
    return data

def plotData(data, plotDataDictArray, showPlot=True, figure=None):
    if figure is None: figure = plt.figure(figsize=(10, 6))

    for index, plotData in enumerate(plotDataDictArray):
        xKey = plotData.get("xKey")
        xKeyFormatter = plotData.get("xKeyFormatter", lambda x: x)
        yKeys = plotData.get("yKeys", [])
        title = plotData.get("title", "Data over " + xKey)
        xLabel = plotData.get("xLabel", xKey)
        yLabel = plotData.get("yLabel", "Data")
        xLimit = plotData.get("xLimit", None)
        yLimit = plotData.get("yLimit", None)

        ax = figure.add_subplot(len(plotDataDictArray), 1, index + 1)
        for yKey in yKeys:  
            try:
                ax.plot([xKeyFormatter(t) for t in data[xKey]], data[yKey], label=yKey)
            except Exception as e:
                print(f"Error plotting key '{yKey}': {e}")
                continue

        ax.set_title(title)
        ax.set_xlabel(xLabel)
        ax.set_ylabel(yLabel)
        if xLimit: ax.set_xlim(xLimit)
        if yLimit: ax.set_ylim(yLimit)
        ax.legend()
        ax.grid(True)

    plt.grid(True)

    if showPlot: plt.show()
